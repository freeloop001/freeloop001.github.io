import openpyxl
import os

# Excel 文件处理
# 需要安装: pip install openpyxl

# 创建 Excel 文件
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "学生成绩"

# 写入数据
ws.append(["姓名", "数学", "语文", "英语"])
ws.append(["张三", 95, 88, 92])
ws.append(["李四", 87, 90, 85])

# 保存文件
wb.save("scores.xlsx")
print("Excel 文件已创建: scores.xlsx")

# 读取 Excel 文件
wb = openpyxl.load_workbook("scores.xlsx")
ws = wb.active
print("读取 Excel 数据:")
for row in ws.iter_rows(values_only=True):
    print(row)

# 清理
if os.path.exists("scores.xlsx"):
    os.remove("scores.xlsx")
